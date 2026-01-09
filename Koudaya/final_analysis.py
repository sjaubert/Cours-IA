import openpyxl
import re

print("="*80)
print("ANALYSE COMPLETE DE LA TRANSFORMATION CSV -> EXCEL")
print("="*80)

# ===== 1. ANALYSE DU CSV =====
with open('251114_101132.csv', 'rb') as f:
    content = f.read().decode('utf-8')

# Le fichier utilise \r comme séparateur de lignes
csv_lines = content.strip().split('\r')
csv_matrix = []
for line_idx, line in enumerate(csv_lines):
    values = [float(v.strip()) for v in line.split(',') if v.strip()]
    csv_matrix.append(values)

print(f"\n* FICHIER CSV:")
print(f"   Structure: {len(csv_matrix)} lignes × {len(csv_matrix[0])} colonnes")
print(f"   Total de valeurs: {sum(len(line) for line in csv_matrix)}")
print(f"\n   Premieres lignes du CSV:")
for i in range(min(3, len(csv_matrix))):
    print(f"   Ligne {i+1}: {csv_matrix[i][:5]}... ({len(csv_matrix[i])} valeurs)")

# ===== 2. ANALYSE DE L'EXCEL =====
wb = openpyxl.load_workbook('Recette_2.xlsx')
ws = wb.active

headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
data_rows = list(ws.iter_rows(min_row=2, values_only=True))

print(f"\n* FICHIER EXCEL:")
print(f"   Feuille: {ws.title}")
print(f"   Structure: {len(data_rows)} lignes de donnees × {len(headers)} colonnes")
print(f"\n   En-tetes:")
print(f"     - Colonnes temporelles: {headers[0]}, {headers[1]}")
print(f"     - Colonnes 'P PMET': {headers[2:12]}")
print(f"     - Colonnes 'G PMET': {headers[12:22]}")

# ===== 3. MAPPER LA TRANSFORMATION =====
print(f"\n* LOGIQUE DE TRANSFORMATION:")
print(f"\n   Le CSV contient des donnees brutes en matrice {len(csv_matrix)}×{len(csv_matrix[0])}")
print(f"   L'Excel ajoute des metadonnees (dates, heures) et reorganise les donnees")
print(f"\n   Observation:")
print(f"   - CSV: {len(csv_matrix)} lignes de mesures")
print(f"   - Excel: {len(data_rows)} lignes de donnees (chaque ligne = 1 mesure temporelle)")
print(f"   - Difference: {len(data_rows)} - {len(csv_matrix)} = {len(data_rows) - len(csv_matrix)} lignes supplementaires")

# Analyser les valeurs pour trouver la correspondance
print(f"\n   > Recherche de correspondance des valeurs:")
print(f"\n   Premiere ligne CSV: {csv_matrix[0][:10]}")
print(f"\n   Valeurs dans l'Excel (colonnes 'P PMET' ligne 2):")
row2_pmet = data_rows[0][2:12]
print(f"   {row2_pmet}")

# Chercher ces valeurs dans le CSV
print(f"\n   * MAPPING DETECTE:")
flatten_csv = []
for line in csv_matrix:
    flatten_csv.extend(line)

# Commençons par voir si les valeurs correspondent dans un ordre particulier
print(f"\n   Toutes les valeurs du CSV (premiere moitie - 'P PMET' potentielles):")
for i in range(min(10, len(csv_matrix[0]))):
    col_values = [csv_matrix[row][i] for row in range(len(csv_matrix)) if i < len(csv_matrix[row])]
    print(f"   Colonne CSV {i+1}: {col_values[:5]}...")
    
print(f"\n   Valeurs 'P PMET' dans l'Excel (premiere colonne de mesure):")
excel_pmet_col1 = [row[2] for row in data_rows[:10]]
print(f"   {excel_pmet_col1}")

# ===== 4. HYPOTHÈSE =====
print(f"\n* HYPOTHESE DE TRANSFORMATION:")
print(f"   L'application semble:")
print(f"   1. Lire la matrice CSV de {len(csv_matrix)}×{len(csv_matrix[0])}")
print(f"   2. Reorganiser les donnees (possiblement transposer ou rearranger)")
print(f"   3. Ajouter des colonnes Date/Heure pour chaque ligne")
print(f"   4. Separer les mesures en deux groupes: 'P PMET' (10 cols) et 'G PMET' (10 cols)")
print(f"   5. Generer {len(data_rows)} lignes avec timestamps")

# Vérifier si c'est une transposition
print(f"\n   Test de transposition:")
if len(csv_matrix) == 15 and len(csv_matrix[0]) == 15:
    print(f"   CSV est une matrice carree 15×15")
    print(f"   Excel a 20 colonnes de mesures (P+G PMET)")
    print(f"   Hypothese: peut-etre 10 colonnes par type × 2 types?")
