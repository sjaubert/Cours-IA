"""
Analyse graphique des mesures PMET
Génère des graphiques de tendance pour comparer les relations entre vis
"""

from colorsys import TWO_THIRD
import pandas as pd
import matplotlib.pyplot as plt
import os
import json
from datetime import datetime
import traceback

# Définition par défaut des relations (pour compatibilité ou cas standard 20 cotes)
DEFAULT_RELATIONS = [
    (1, 5, "Relation 1"),  # VIS 1 avec VIS 5
    (2, 6, "Relation 2"),  # VIS 2 avec VIS 6
    (3, 7, "Relation 3"),  # VIS 3 avec VIS 7
    (4, 10, "Relation 4"), # VIS 4 avec VIS 10
    (8, 9, "Relation 5"),  # VIS 8 avec VIS 9
]

def get_dynamic_relations(df, pmet_type):
    """
    Détermine dynamiquement les relations en fonction du nombre de cotes disponibles.
    Règle: 10 relations pour 20 cotes (5 par type), 8 relations pour 18 cotes (4 par type).
    """
    prefix = f'{pmet_type} PMET Cote vis '
    available_vis = []
    for i in range(1, 21): 
        if f'{prefix}{i}' in df.columns:
            available_vis.append(i)
    
    num_vis = len(available_vis)
    print(f"Détection dynamique pour {pmet_type} PMET: {num_vis} vis trouvées.")
    
    relations = []
    
    # Règle demandée :
    # Si 20 cotes au total -> 10 relations (5 P + 5 G). 10 vis par type -> 5 relations.
    # Si 18 cotes au total -> 8 relations (4 P + 4 G). 9 vis par type -> 4 relations.
    # On apparie (i, i + num_vis/2)
    
    num_pairs = num_vis // 2
    
    for i in range(1, num_pairs + 1):
        v1 = i
        v2 = i + num_pairs
        if f'{prefix}{v1}' in df.columns and f'{prefix}{v2}' in df.columns:
            relations.append((v1, v2, f"Relation {len(relations)+1}"))
            
    return relations

def load_data(file_path):
    """Charge les données depuis le fichier Excel"""
    print(f"Chargement des données depuis {file_path}...")
    try:
        df = pd.read_excel(file_path)
        print(f"Données chargées: {df.shape[0]} mesures, {df.shape[1]} colonnes")
        return df
    except Exception as e:
        print(f"Erreur lors du chargement du fichier: {e}")
        raise

def create_output_directory(output_dir):
    """Crée le dossier de sortie s'il n'existe pas"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Dossier '{output_dir}' créé")

def plot_relations(df, pmet_type, output_dir):
    """
    Génère les graphiques de tendance pour un type PMET (P ou G)
    """
    relations = get_dynamic_relations(df, pmet_type)
    num_relations = len(relations)
    
    if num_relations == 0:
        print(f"Aucune relation possible pour {pmet_type} PMET.")
        return None

    # Créer un numéro de moule séquentiel (1 à n)
    num_moules = range(1, len(df) + 1)
    
    # L'offset dépend du nombre de relations du type précédent (si on veut numéroter en continu)
    # Mais ici on va faire plus simple: P = 1..N, G = (N+1)..2N
    # Pour 20 cotes (5 rel P, 5 rel G), P=1..5, G=6..10.
    # Pour 18 cotes (4 rel P, 4 rel G), P=1..4, G=5..8.
    
    # On calcule l'offset dynamiquement
    if pmet_type == 'P':
        offset = 0
    else:
        # On recalcule les relations P pour connaître l'offset
        p_relations = get_dynamic_relations(df, 'P')
        offset = len(p_relations)
        
    limite_qualite = 80 if pmet_type == 'P' else 130
    
    # Créer la figure avec N sous-graphiques
    fig, axes = plt.subplots(num_relations, 1, figsize=(14, 3.5 * num_relations))
    # Si un seul graphique, axes n'est pas une liste
    if num_relations == 1:
        axes = [axes]
        
    fig.suptitle(f'{pmet_type} PMET - Analyse des Relations (Relations {offset+1} à {offset+num_relations})', 
                 fontsize=16, fontweight='bold', y=0.995)
    
    # Calculer les statistiques GLOBALES
    all_cols = [f'{pmet_type} PMET Cote vis {i}' for i in range(1, 11)]
    valid_cols = [c for c in all_cols if c in df.columns]
    
    result_stats = None

    if valid_cols:
        nb_moules = len(df)
        total_vis = len(valid_cols) * nb_moules
        # Nb DEFAUT PMET (Total mesures < limite)
        all_values = df[valid_cols].values.flatten()
        nb_defaut_pmet = (all_values < limite_qualite).sum()
        # Nb DEFAUT MOULE
        moules_defectueux = (df[valid_cols] < limite_qualite).any(axis=1).sum()
        # Pourcentages
        non_conforme_pct = (moules_defectueux / nb_moules) * 100 if nb_moules > 0 else 0
        pmet_defaut_pct = (nb_defaut_pmet / total_vis) * 100 if total_vis > 0 else 0
        global_mean = all_values.mean()
        
        # Texte du résumé global
        type_label = f"{pmet_type}PMET"
        global_stats_text = (
            f"MOULE: {nb_moules}\n"
            f"Nb DEFAUT MOULE: {moules_defectueux}\n"
            f"Nb DEFAUT PMET: {nb_defaut_pmet}\n"
            f"NON CONFORME%: {non_conforme_pct:.2f}%\n"
            f"{type_label} DEFAUT%: {pmet_defaut_pct:.2f}%\n"
            f"MOYENNE: {global_mean:.2f}"
        )
        
        props = dict(boxstyle='round', facecolor='azure', alpha=0.9, linewidth=1, edgecolor='blue')
        fig.text(0.82, 0.98, global_stats_text, fontsize=9, 
                 verticalalignment='top', horizontalalignment='left', bbox=props)
        
        result_stats = {
            "type": pmet_type,
            "total_moules": int(nb_moules),
            "moules_defectueux": int(moules_defectueux),
            "total_defauts_vis": int(nb_defaut_pmet),
            "taux_non_conformite": float(non_conforme_pct),
            "taux_defaut_vis": float(pmet_defaut_pct),
            "moyenne_globale": float(global_mean),
            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M")
        }

    # Pour chaque relation
    for idx, (vis1, vis2, _) in enumerate(relations):
        ax = axes[idx]
        relation_num = idx + 1 + offset
        col_vis1 = f'{pmet_type} PMET Cote vis {vis1}'
        col_vis2 = f'{pmet_type} PMET Cote vis {vis2}'
        
        # Vérifier si les colonnes existent
        if col_vis1 not in df.columns or col_vis2 not in df.columns:
            ax.text(0.5, 0.5, "Données manquantes pour cette relation", 
                    ha='center', va='center', transform=ax.transAxes)
            continue

        ax.axhline(y=limite_qualite, color='red', linestyle='--', linewidth=2, 
                   label=f'Limite min ({limite_qualite})', alpha=0.7, zorder=1)
        
        ax.plot(num_moules, df[col_vis1], marker='o', markersize=4, 
                linewidth=1.5, label=f'VIS {vis1}', alpha=0.8, zorder=2)
        ax.plot(num_moules, df[col_vis2], marker='s', markersize=4, 
                linewidth=1.5, label=f'VIS {vis2}', alpha=0.8, zorder=2)
        
        relation_mean = df[[col_vis1, col_vis2]].values.mean()
        mean_text = f"Moyenne Relation: {relation_mean:.2f} mm"
        props_relation = dict(boxstyle='round', facecolor='white', alpha=0.8, linewidth=1, edgecolor='gray')
        ax.text(0.02, 0.95, mean_text, transform=ax.transAxes, fontsize=9,
                verticalalignment='top', horizontalalignment='left', bbox=props_relation)
        
        ax.set_title(f'Relation {relation_num}: VIS {vis1} ↔ VIS {vis2}', 
                    fontsize=12, fontweight='bold', pad=10)
        ax.set_xlabel('Numéro de moule', fontsize=10)
        ax.set_ylabel('Mesure (mm)', fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.legend(loc='upper right', fontsize=8) 
    
    plt.tight_layout(rect=[0, 0, 1, 0.94])
    
    output_file = os.path.join(output_dir, f'{pmet_type}_PMET_tendances.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    return result_stats

def analyze_file(input_file, output_dir_base):
    """Fonction principale encapsulée pour appel externe"""
    print("\n" + "="*70)
    print("ANALYSE GRAPHIQUE DES MESURES PMET")
    print("="*70 + "\n")
    
    # 1. Charger les données
    df = load_data(input_file)
    
    # 2. Préparer les dossiers
    # Dossier pour les graphiques
    graph_dir = os.path.join(output_dir_base, 'graphiques')
    create_output_directory(graph_dir)
    
    all_stats = {}
    
    # 3. Générer les graphiques P PMET
    print("\nGénération des graphiques P PMET...")
    all_stats['P'] = plot_relations(df, 'P', graph_dir)
    
    # 4. Générer les graphiques G PMET
    print("Génération des graphiques G PMET...")
    all_stats['G'] = plot_relations(df, 'G', graph_dir)
    
    # 5. Sauvegarder stats
    # summary dir est dans output_dir_base/summary
    summary_dir = os.path.join(output_dir_base, 'summary')
    if not os.path.exists(summary_dir):
        os.makedirs(summary_dir)
        
    stats_file = os.path.join(summary_dir, 'stats.json')
    with open(stats_file, 'w', encoding='utf-8') as f:
        json.dump(all_stats, f, indent=4, ensure_ascii=False)
        
    print(f"Statistiques sauvegardées: {stats_file}")

    # 6. Exporter vers Excel (Simple)
    excel_file = export_stats_to_excel(all_stats, output_dir_base)
    
    # 7. Exporter vers Excel (Interactif)
    rich_excel_file = export_rich_excel(all_stats, output_dir_base, df)
    
    print("Analyse terminée.")
    return all_stats, graph_dir, excel_file, rich_excel_file

def export_stats_to_excel(all_stats, output_dir):
    """
    Génère un rapport Excel simple résumant les statistiques P et G PMET.
    """
    rows = []
    
    for pmet_type, data in all_stats.items():
        if not data: continue
        
        row = {
            "Type PMET": f"{pmet_type} PMET",
            "Moyenne Globale (mm)": data.get("moyenne_globale"),
            "Total Moules": data.get("total_moules"),
            "Moules Défectueux": data.get("moules_defectueux"),
            "Total Défauts Vis": data.get("total_defauts_vis"),
            "Taux Non-Conformité (%)": data.get("taux_non_conformite"),
            "Taux Défaut Vis (%)": data.get("taux_defaut_vis"),
            "Date Analyse": data.get("timestamp")
        }
        rows.append(row)
        
    if not rows:
        return None
        
    df_export = pd.DataFrame(rows)
    output_file = os.path.join(output_dir, "Rapport_Analyse_PMET_Simple.xlsx")
    
    try:
        df_export.to_excel(output_file, index=False, sheet_name="Synthèse")
        return output_file
    except Exception as e:
        print(f"Erreur lors de l'export Excel simple: {e}")
        return None

def export_rich_excel(all_stats, output_dir, df_input):
    """
    Génère un rapport Excel COMPLET avec des GRAPHIQUES NATIFS liés aux données.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        print("Erreur: openpyxl non installé.")
        return None

    output_file = os.path.join(output_dir, "Rapport_Complet_PMET.xlsx")

    try:
        wb = Workbook()
        ws_synth = wb.active
        ws_synth.title = "Synthèse Complète"
        
        # --- STYLES ---
        title_font = Font(name='Arial', size=20, bold=True, color='333333')
        subtitle_font = Font(name='Arial', size=14, bold=True, color='0066CC')
        header_font = Font(name='Arial', size=11, bold=True)
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        border_style = Side(border_style='thin')
        box_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # En-tête Global
        ws_synth['B2'] = 'SATEBA France - Rapport d\'Analyse PMET (Interactif)'
        ws_synth['B2'].font = title_font
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws_synth['B3'] = f'Généré le: {timestamp}'
        ws_synth['B3'].font = Font(size=10, italic=True, color='666666')
        
        current_row = 5

        # --- CRÉATION DES ONGLETS DE DONNÉES ---
        def create_data_sheet(pmet_type):
            sheet_name = f"Données {pmet_type}"
            ws_data = wb.create_sheet(sheet_name)
            
            all_cols = [f'{pmet_type} PMET Cote vis {i}' for i in range(1, 11)]
            cols_to_keep = [c for c in all_cols if c in df_input.columns]
            
            # Header: MOULE count + cotes
            ws_data.cell(row=1, column=1, value="MOULE")
            for i, col in enumerate(cols_to_keep, start=2):
                ws_data.cell(row=1, column=i, value=col)
            
            # Data
            for r_idx, (idx, row) in enumerate(df_input.iterrows(), start=2):
                ws_data.cell(row=r_idx, column=1, value=r_idx - 1) # Numéro de moule 1..n
                for c_idx, col in enumerate(cols_to_keep, start=2):
                    ws_data.cell(row=r_idx, column=c_idx, value=row[col])
            
            return ws_data, cols_to_keep

        # --- FONCTION POUR AJOUTER GRAPHIQUE ---
        def add_native_chart(ws_synth, ws_data, data_cols, start_row, pmet_type, relations):
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.colors import ColorChoice
            from openpyxl.chart.axis import ChartLines

            row_cursor = start_row
            num_moules = len(df_input)
            limite_qualite = 80 if pmet_type == 'P' else 130
            
            # Pour chaque relation, on crée un graphique
            for i_rel, (vis1, vis2, _) in enumerate(relations):
                col_name1 = f'{pmet_type} PMET Cote vis {vis1}'
                col_name2 = f'{pmet_type} PMET Cote vis {vis2}'
                
                if col_name1 not in data_cols or col_name2 not in data_cols:
                    continue
                
                idx1 = data_cols.index(col_name1) + 2
                idx2 = data_cols.index(col_name2) + 2
                
                # Calcul de l'offset pour la numérotation
                if pmet_type == 'P':
                    rel_num = i_rel + 1
                else:
                    p_rels = get_dynamic_relations(df_input, 'P')
                    rel_num = i_rel + 1 + len(p_rels)

                # Calcul de la moyenne
                mean_val = (df_input[col_name1].mean() + df_input[col_name2].mean()) / 2
                
                chart = LineChart()
                chart.title = f"Relation {rel_num}: VIS {vis1} ↔ VIS {vis2}"
                chart.style = 13
                # Ajustement des axes pour que ce soit clair (similaire au web)
                chart.y_axis.title = 'Mesure (mm)'
                chart.x_axis.title = 'Numéro de moule'
                chart.height = 12
                chart.width = 25
                chart.legend.position = 'b'
                
                # Grilles plus discrètes
                chart.y_axis.majorGridlines = ChartLines()
                chart.x_axis.majorGridlines = None
                
                # Ajustement des axes pour que ce soit clair (similaire au web)
                # On définit un min/max intelligent
                all_vals = df_input[[col_name1, col_name2]].values.flatten()
                min_data = all_vals.min()
                max_data = all_vals.max()
                
                # Laisser de la place pour la limite rouge
                y_min = min(min_data, limite_qualite) - 5
                y_max = max(max_data, limite_qualite) + 5
                
                chart.y_axis.scaling.min = max(0, float(y_min))
                chart.y_axis.scaling.max = float(y_max)

                # Data series
                data1 = Reference(ws_data, min_col=idx1, min_row=1, max_row=num_moules+1)
                data2 = Reference(ws_data, min_col=idx2, min_row=1, max_row=num_moules+1)
                
                chart.add_data(data1, titles_from_data=True)
                chart.add_data(data2, titles_from_data=True)
                
                # Couleurs (Blue and Orange like Web version)
                # Série 1: Bleu (Web style)
                s1 = chart.series[0]
                s1.graphicalProperties.line.solidFill = "1F77B4" 
                s1.marker.symbol = "circle"
                s1.marker.size = 4
                s1.graphicalProperties.line.width = 15000 # Environ 1.5pt
                
                # Série 2: Orange (Web style)
                s2 = chart.series[1]
                s2.graphicalProperties.line.solidFill = "FF7F0E"
                s2.marker.symbol = "square"
                s2.marker.size = 4
                s2.graphicalProperties.line.width = 15000

                # --- AJOUT DE LA LIGNE DE LIMITE ROUGE ---
                # On crée une série fictive pour la limite
                limit_col_idx = len(data_cols) + 3 # Une colonne après les données
                ws_data.cell(row=1, column=limit_col_idx, value=f"Limite ({limite_qualite})")
                for r in range(2, num_moules + 2):
                    ws_data.cell(row=r, column=limit_col_idx, value=limite_qualite)
                
                limit_ref = Reference(ws_data, min_col=limit_col_idx, min_row=1, max_row=num_moules+1)
                chart.add_data(limit_ref, titles_from_data=True)
                
                s_limit = chart.series[2]
                s_limit.graphicalProperties.line.solidFill = "FF0000" # Rouge
                s_limit.graphicalProperties.line.dashStyle = "dash"
                s_limit.graphicalProperties.line.width = 25000 # Plus épais
                s_limit.marker.symbol = "none"

                # X-axis labels
                cats = Reference(ws_data, min_col=1, min_row=2, max_row=num_moules+1)
                chart.set_categories(cats)
                
                # Affichage de la moyenne (Styled like a badge)
                cell_avg = ws_synth.cell(row=row_cursor, column=2, value=f"MOYENNE RELATION {rel_num} : {mean_val:.2f} mm")
                cell_avg.font = Font(name='Arial', size=11, bold=True, color='0066CC')
                cell_avg.alignment = Alignment(horizontal='left')
                row_cursor += 1
                
                # Ajouter à la synthèse
                anchor = f"B{row_cursor}"
                ws_synth.add_chart(chart, anchor)
                row_cursor += 25
            
            return row_cursor

        # --- EXÉCUTION DES SECTIONS ---
        for p_type in ['P', 'G']:
            if p_type not in all_stats or not all_stats[p_type]: continue
            
            data = all_stats[p_type]
            ws_data, cols_kept = create_data_sheet(p_type)
            
            # Titre Section avec le range de relations
            p_rels = get_dynamic_relations(df_input, 'P')
            num_p = len(p_rels)
            if p_type == 'P':
                range_text = f"Relations 1 à {num_p}"
            else:
                g_rels = get_dynamic_relations(df_input, 'G')
                range_text = f"Relations {num_p + 1} à {num_p + len(g_rels)}"

            section_title = f'ANALYSE {p_type} PMET - {range_text} ({"< 80mm" if p_type=="P" else "< 130mm"})'
            ws_synth.cell(row=current_row, column=2, value=section_title).font = subtitle_font
            current_row += 2
            
            # Tableau des Stats
            headers = ['MOULE', 'Nb DEFAUT MOULE', 'Nb DEFAUT PMET', 'NON CONFORME%', f'{p_type}PMET DEFAUT%', 'MOYENNE']
            values = [
                data.get("total_moules", 0),
                data.get("moules_defectueux", 0),
                data.get("total_defauts_vis", 0),
                f"{data.get('taux_non_conformite', 0):.2f}%", 
                f"{data.get('taux_defaut_vis', 0):.2f}%",
                f"{data.get('moyenne_globale', 0):.2f}"
            ]
            
            for i, (h, v) in enumerate(zip(headers, values)):
                c_h = ws_synth.cell(row=current_row, column=i+2, value=h)
                c_h.font = header_font
                c_h.fill = header_fill
                c_h.border = box_border
                c_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                c_v = ws_synth.cell(row=current_row+1, column=i+2, value=v)
                c_v.border = box_border
                c_v.alignment = Alignment(horizontal='center')
            
            current_row += 4
            
            # Graphiques Natifs
            relations = get_dynamic_relations(df_input, p_type)
            current_row = add_native_chart(ws_synth, ws_data, cols_kept, current_row, p_type, relations)
            current_row += 3

        ws_synth.column_dimensions['B'].width = 25
        ws_synth.column_dimensions['C'].width = 25
        ws_synth.column_dimensions['D'].width = 25
        ws_synth.column_dimensions['E'].width = 25
        
        wb.save(output_file)
        return output_file

    except Exception as e:
        print(f"Erreur lors de l'export Excel interactif: {e}")
        traceback.print_exc()
        return None

# Point d'entrée pour exécution directe (héritage)
if __name__ == "__main__":
    # Valeurs par défaut pour compatibilité
    DEFAULT_FILE = 'Recette_2.xlsx'
    DEFAULT_DIR = os.getcwd()
    if os.path.exists(DEFAULT_FILE):
        analyze_file(DEFAULT_FILE, DEFAULT_DIR)
    else:
        print(f"Fichier par défaut {DEFAULT_FILE} non trouvé.")

