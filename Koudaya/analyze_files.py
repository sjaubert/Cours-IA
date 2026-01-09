import openpyxl

print("="*60)
print("ANALYSE DU FICHIER CSV")
print("="*60)

# Lire le CSV
with open('251114_101132.csv', 'r') as f:
    content = f.read()

# Le fichier semble utiliser \r comme séparateur de lignes
lines = content.strip().split('\r')
print(f"\nNombre de lignes (séparées par \\r): {len(lines)}")

# Analyser chaque ligne
csv_data = []
for i, line in enumerate(lines):
    values = [v.strip() for v in line.split(',') if v.strip()]
    csv_data.append(values)
    if i < 3:
        print(f"\nLigne {i+1}: {len(values)} valeurs")
        print(f"  Premières valeurs: {values[:5]}")

print(f"\nTotal lignes CSV: {len(csv_data)}")
print(f"Valeurs par ligne: {len(csv_data[0]) if csv_data else 0}")

print("\n" + "="*60)
print("ANALYSE DU FICHIER EXCEL")
print("="*60)

wb = openpyxl.load_workbook('Recette_2.xlsx')
print(f"\nFeuilles disponibles: {wb.sheetnames}")

ws = wb.active
print(f"Feuille active: {ws.title}")
print(f"Dimensions: {ws.dimensions}")

print("\nPremières lignes de l'Excel:")
excel_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    excel_data.append(row)
    if i < 10:
        print(f"Ligne {i+1}: {row}")

print(f"\nTotal lignes Excel: {len(excel_data)}")

print("\n" + "="*60)
print("COMPARAISON")
print("="*60)
print(f"CSV: {len(csv_data)} lignes x {len(csv_data[0])} colonnes")
print(f"Excel: {len(excel_data)} lignes x {len(excel_data[0]) if excel_data else 0} colonnes")
