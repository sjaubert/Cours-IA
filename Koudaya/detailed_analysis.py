import openpyxl

print("="*80)
print("ANALYSE DÉTAILLÉE DE LA TRANSFORMATION")
print("="*80)

# Lire le CSV
with open('251114_101132.csv', 'r') as f:
    content = f.read()

# Remplacer tous les retours à la ligne par rien pour avoir une seule ligne
content = content.replace('\r', '').replace('\n', '')
# Séparer par virgules
values_str = [v.strip() for v in content.split(',') if v.strip()]
all_values = [float(v) for v in values_str]

print(f"\n1. FICHIER CSV:")
print(f"   - Total de valeurs: {len(all_values)}")
print(f"   - Premières 30 valeurs:")
for i in range(0, min(30, len(all_values)), 10):
    print(f"     [{i:3d}-{i+9:3d}]: {all_values[i:i+10]}")

# Lire l'Excel
wb = openpyxl.load_workbook('Recette_2.xlsx')
ws = wb.active

print(f"\n2. FICHIER EXCEL:")
print(f"   - Dimensions: {ws.dimensions}")
print(f"   - En-têtes (ligne 1):")
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
for i, h in enumerate(headers):
    print(f"     Col {i+1:2d}: {h}")

print(f"\n3. STRUCTURE DES DONNÉES:")
print(f"   - L'Excel a {len(headers)} colonnes dont:")
print(f"     * 2 colonnes temporelles (Date, Heure)")
print(f"     * {len(headers)-2} colonnes de mesures")
print(f"   - Nombre de colonnes de mesures: {len(headers)-2}")
print(f"   - Si on divise 225 valeurs par 15 = {225/15}")

# Analyser les valeurs des premières lignes de données
print(f"\n4. PREMIÈRES LIGNES DE DONNÉES (sans en-tête):")
data_rows = list(ws.iter_rows(min_row=2, max_row=5, values_only=True))
for i, row in enumerate(data_rows):
    print(f"\n   Ligne {i+2}:")
    print(f"     Date: {row[0]}, Heure: {row[1]}")
    print(f"     Valeurs mesures: {row[2:7]}... (20 mesures au total)")

print(f"\n5. HYPOTHÈSE DE TRANSFORMATION:")
print(f"   - CSV: 225 valeurs = 15 lignes × 15 valeurs")
print(f"   - Excel: Chaque ligne semble avoir 20 mesures (10 'P PMET' + 10 'G PMET')")
print(f"   - Il semble y avoir une réorganisation matricielle des données")

# Vérifier le pattern dans le CSV
print(f"\n6. PATTERN DES VALEURS CSV:")
print(f"   Les 15 premières valeurs:")
for i in range(15):
    print(f"     Val {i+1:2d}: {all_values[i]:8.3f}")
print(f"   Les 15 suivantes (valeurs 16-30):")
for i in range(15, 30):
    print(f"     Val {i+1:2d}: {all_values[i]:8.3f}")

# Comparer avec les valeurs Excel
print(f"\n7. COMPARAISON CSV -> EXCEL (ligne 2 de l'Excel):")
row2 = data_rows[0]
print(f"   Excel ligne 2, colonnes P PMET (3-12): {row2[2:12]}")
print(f"   Excel ligne 2, colonnes G PMET (13-22): {row2[12:22]}")
