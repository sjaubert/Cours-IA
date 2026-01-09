import openpyxl
import pandas as pd

print("=== Analyse du fichier Excel ===")
wb = openpyxl.load_workbook('Recette_2.xlsx')
print(f"Feuilles: {wb.sheetnames}")

ws = wb.active
print(f"Feuille active: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print("\n=== Premières lignes ===")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(f"Ligne {i+1}: {row}")
    if i >= 20:
        break

print("\n=== Analyse du fichier CSV ===")
with open('251114_101132.csv', 'r') as f:
    content = f.read()
    lines = content.split('\r')
    print(f"Nombre de lignes: {len(lines)}")
    print(f"Nombre de valeurs par ligne: {len(lines[0].split(',')) if lines else 0}")
    print("\nPremières lignes du CSV:")
    for i, line in enumerate(lines[:5]):
        values = line.split(',')
        print(f"Ligne {i+1}: {len(values)} valeurs - {values[:5]}...")
